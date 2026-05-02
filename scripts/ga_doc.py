import argparse
import os
import subprocess
import sys
import zipfile

sys.path.insert(0, os.path.dirname(__file__))
from _common import *


def read_docx(path):
    """使用 pandoc 将 docx 转纯文本"""
    r = subprocess.run(
        ["pandoc", path, "-t", "plain", "--wrap=none"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="ignore",
    )
    if r.returncode == 0:
        return r.stdout

    from xml.etree import ElementTree as ET

    with zipfile.ZipFile(path) as z:
        with z.open("word/document.xml") as f:
            tree = ET.parse(f)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    texts = []
    for p in tree.findall(".//w:p", ns):
        para_text = "".join(t.text or "" for t in p.findall(".//w:t", ns))
        texts.append(para_text)
    return "\n".join(texts)


def read_pdf(path):
    import pdfplumber

    texts = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            t = page.extract_text()
            if t:
                texts.append(f"--- Page {i + 1} ---\n{t}")
    return "\n\n".join(texts)


def read_xlsx(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    output = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(max_row=50, values_only=True):
            output.append("\t".join(str(c) if c is not None else "" for c in row))
    return "\n".join(output)


@confirm_required
def cmd_images(args, dry_run=True):
    if not os.path.exists(args.file):
        fail(f"文件不存在: {args.file}")
        return EXIT_FAIL

    if dry_run:
        info(f"[预览] 将从 {args.file} 提取图片到 {args.out}")
        return EXIT_SKIP

    os.makedirs(args.out, exist_ok=True)
    count = 0
    with zipfile.ZipFile(args.file) as z:
        for name in z.namelist():
            if name.startswith("word/media/"):
                z.extract(name, args.out)
                count += 1
    ok(f"提取了 {count} 张图片到 {args.out}/word/media/")
    return EXIT_OK


@confirm_required
def cmd_merge(args, dry_run=True):
    if len(args.files) < 2:
        fail("merge 至少需要两个 PDF")
        return EXIT_FAIL

    missing = [f for f in args.files if not os.path.exists(f)]
    if missing:
        fail(f"文件不存在: {', '.join(missing)}")
        return EXIT_FAIL

    if dry_run:
        info(f"[预览] 将合并 {len(args.files)} 个 PDF -> {args.out}")
        return EXIT_SKIP

    from pypdf import PdfMerger

    merger = PdfMerger()
    for f in args.files:
        merger.append(f)
    merger.write(args.out)
    merger.close()
    ok(f"合并 {len(args.files)} 个 PDF -> {args.out}")
    return EXIT_OK


def cmd_read(args):
    if not os.path.exists(args.file):
        fail(f"文件不存在: {args.file}")
        return EXIT_FAIL

    ext = os.path.splitext(args.file)[1].lower()
    try:
        if ext == ".docx":
            result = read_docx(args.file)
        elif ext == ".pdf":
            result = read_pdf(args.file)
        elif ext in (".xlsx", ".xls"):
            result = read_xlsx(args.file)
        else:
            fail(f"不支持的格式: {ext}")
            return EXIT_FAIL
    except Exception as exc:
        fail(f"读取失败: {exc}")
        return EXIT_FAIL

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(result)
        ok(f"已写入: {args.out}")
    else:
        print(result)
    return EXIT_OK


def main():
    parser = argparse.ArgumentParser(description="GA 文档处理")
    sub = parser.add_subparsers(dest="command")

    p_read = sub.add_parser("read")
    p_read.add_argument("file")
    p_read.add_argument("--out", help="输出到文件")

    p_images = sub.add_parser("images")
    p_images.add_argument("file")
    p_images.add_argument("--out", default="./extracted_images")
    p_images.add_argument("--confirm", action="store_true")

    p_merge = sub.add_parser("merge")
    p_merge.add_argument("files", nargs="+")
    p_merge.add_argument("--out", required=True)
    p_merge.add_argument("--confirm", action="store_true")

    args = parser.parse_args()

    if args.command == "read":
        sys.exit(cmd_read(args))
    if args.command == "images":
        sys.exit(cmd_images(args))
    if args.command == "merge":
        sys.exit(cmd_merge(args))

    parser.print_help()
    sys.exit(EXIT_SKIP)


if __name__ == "__main__":
    main()
